# app.py
"""
Final clean version of the Nearest Location Finder app.
- Sidebar: App (User) / Admin
- Admin: upload backend CSV/XLSX, set feasible distance, set backend-duplicate suffix
- User: upload/paste input, parse, choose lat/lon if needed, apply filters, choose Nth (global or per-row via input column),
        compute nearest matching row(s), output combined rows (input columns unchanged; backend columns appended;
        backend columns that conflict with input columns get a suffix configured by Admin).
No external distance libraries required; uses numpy/pandas only.
"""

import streamlit as st
import pandas as pd
import numpy as np
import math
from io import StringIO
import os
import json
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter


#st.write("DEBUG - Secrets available:", dict(st.secrets))
# --------------------------
# Simple Authentication
# --------------------------
# -------------------------
# Admin login helper (using secrets)
# -------------------------
def admin_login():
    st.subheader("Admin Login")

    username_input = st.sidebar.text_input("Username")
    password_input = st.sidebar.text_input("Password", type="password")

    # Fetch credentials from secrets
    #admin_user = st.secrets["admin"]["username"]
    #admin_pass = st.secrets["admin"]["password"]

    if st.sidebar.button("Login"):
        if username_input == "admin" and password_input == "1234":
            st.session_state["admin_authenticated"] = True
            st.success("✅ Logged in as admin")
            st.rerun()
        else:
            st.error("❌ Invalid username or password")
# -------------------------
# User login helper
# -------------------------
def user_login():
    st.subheader("User Login")

    username_input = st.sidebar.text_input("Username", key="user_username")
    password_input = st.sidebar.text_input("Password", type="password", key="user_password")

    if st.sidebar.button("Login", key="user_login_button"):
        # Hardcoded credentials
        if username_input == "user" and password_input == "5678":
            st.session_state["user_authenticated"] = True
            st.success("✅ Logged in as user")
            st.rerun()
        else:
            st.error("❌ Invalid username or password")


# Initialize session state
if "admin_authenticated" not in st.session_state:
    st.session_state["admin_authenticated"] = False

if "user_authenticated" not in st.session_state:
    st.session_state["user_authenticated"] = False


# -------------------------
# Config / filenames
# -------------------------
CONFIG_FILE = "app_config.json"
BACKEND_FILE = "backend_data.csv"
DEFAULT_CONFIG = {
    "feasible_km": 20.0,
    "backend_conflict_suffix": "_matched"
}

# Load / save config
def load_config():
    if os.path.exists(CONFIG_FILE):
        try:
            with open(CONFIG_FILE, "r") as f:
                return json.load(f)
        except Exception:
            return DEFAULT_CONFIG.copy()
    return DEFAULT_CONFIG.copy()

def save_config(cfg):
    with open(CONFIG_FILE, "w") as f:
        json.dump(cfg, f)

cfg = load_config()

# -------------------------
# Small helpers
# -------------------------
def safe_read_table(obj, filename=None):
    """Try to read CSV/XLSX from file-like or path."""
    name = (getattr(obj, "name", None) or filename or "").lower()
    try:
        if name.endswith((".xlsx", ".xls")):
            # Force Excel engine to avoid errors in cloud
            return pd.read_excel(obj, engine="openpyxl")
    except Exception as e:
        st.error(f"Excel read failed: {e}")
    try:
        return pd.read_csv(obj, sep=None, engine="python", skipinitialspace=True)
    except Exception as e:
        st.error(f"CSV read failed: {e}")
        raise

def normalize_latlon_names(df):
    """
    Rename many variants to 'Latitude' and 'Longitude'.
    Handles 'Latitute' typo and other common variants.
    """
    col_map = {}
    for col in df.columns:
        low = col.strip().lower()
        # latitude tokens (try aggressive matching but avoid false positives)
        if ("latitute" in low) or ("latitude" in low) or low == "lat" or (low.startswith("lat") and not low.startswith("platform")):
            col_map[col] = "Latitude"
            continue
        # longitude tokens
        if ("longitude" in low) or low in ("lon", "lng", "long") or low.startswith("lon") or low.startswith("long"):
            col_map[col] = "Longitude"
            continue
    if col_map:
        df = df.rename(columns=col_map)
    return df

def detect_latlon_candidates(df):
    """Return best-guess column names (or None) for lat/lon in this dataframe."""
    lat, lon = None, None
    for c in df.columns:
        k = "".join(ch.lower() for ch in str(c).strip() if ch.isalnum())
        if lat is None and k in ("lat", "latitude", "latitute", "latdeg", "latit"):
            lat = c
        if lon is None and k in ("lon", "longitude", "lng", "long", "longit"):
            lon = c
    # fallback: contains 'lat' or 'lon'
    if lat is None:
        for c in df.columns:
            if "lat" in c.lower() and "latitude" not in c.lower():
                lat = c
                break
    if lon is None:
        for c in df.columns:
            if "lon" in c.lower() or "lng" in c.lower():
                lon = c
                break
    return lat, lon

def vectorized_haversine(lat1, lon1, backend_lat_rad, backend_lon_rad):
    """Return distances (km) from (lat1,lon1) to arrays provided in radians."""
    R = 6371.0
    lat1r = math.radians(lat1)
    lon1r = math.radians(lon1)
    dlat = backend_lat_rad - lat1r
    dlon = backend_lon_rad - lon1r
    a = np.sin(dlat/2.0)**2 + np.cos(lat1r)*np.cos(backend_lat_rad)*np.sin(dlon/2.0)**2
    c = 2.0 * np.arctan2(np.sqrt(a), np.sqrt(1.0 - a))
    return R * c

def find_nth_index(distances, n):
    """Given numpy distances vector, return index of nth nearest (1-based n)."""
    if len(distances) == 0:
        return None
    order = np.argsort(distances)
    if n <= 0:
        n = 1
    if n > len(order):
        return None
    return int(order[n-1])

def backend_column_merge_dict(backend_row, input_cols, suffix):
    """
    Return mapping for backend cols -> output column names.
    - if backend column name not in input_cols: keep as-is
    - if conflict: backend col becomes col + suffix
    """
    mapping = {}
    for c in backend_row.index:
        if c in input_cols:
            mapping[c] = f"{c}{suffix}"
        else:
            mapping[c] = c
    return mapping

# -------------------------
# Streamlit UI
# -------------------------
st.set_page_config(page_title="Nearest Site Finder", layout="wide")
st.title("Nearest Site Finder")

# Sidebar navigation
page = st.sidebar.radio("Navigation", ["App", "Admin"])

# -------------------------
# Admin page
# -------------------------
if page == "Admin":
    st.header("Admin — Backend & Settings")

    # Check login state
    if not st.session_state["admin_authenticated"]:
        # Show only login form
        admin_login()
    else:
        # Everything inside here is only visible to logged-in admins
        st.success("✅ Welcome, Admin!")

        # --------------------------
        # Admin-only controls
        # --------------------------
        st.subheader("Upload Backend File")
        uploaded_file = st.file_uploader("Upload CSV", type=["csv"])
        if uploaded_file is not None:
            try:
                df = pd.read_csv(uploaded_file)
                st.session_state["backend_df"] = df
                st.success("✅ Backend file uploaded and stored in memory!")

                st.subheader("🔎 Preview Uploaded Data")
                st.dataframe(df.head())
            except Exception as e:
                st.error(f"Error uploading file: {e}")

        # Show current backend preview
        if os.path.exists(BACKEND_FILE):
            try:
                bdf = safe_read_table(BACKEND_FILE, filename=BACKEND_FILE)
                bdf = normalize_latlon_names(bdf)
                st.subheader("Current backend preview (first 10 rows)")
                st.dataframe(bdf.head(10))
                # st.write(f"Columns: {list(bdf.columns)}")
            except Exception as e:
                st.error(f"Failed to read backend file: {e}")
        else:
            st.info("No backend file found. Upload above.")

        st.markdown("---")

        # Config: feasible distance and suffix for backend conflict
        st.subheader("Settings")
        new_feasible = st.number_input(
        "Feasible distance (km)",
        min_value=1,  # start at 1 km instead of 0.1
        value=int(cfg.get("feasible_km", 20)),  # use int instead of float
        step=1  # ensures whole-number increments
        )
        new_suffix = st.text_input(
        "Suffix to add to backend columns that conflict with input names",
        value=cfg.get("backend_conflict_suffix", "_matched")
        )
        if st.button("Save settings"):
            cfg["feasible_km"] = float(new_feasible)
            cfg["backend_conflict_suffix"] = new_suffix.strip() or "_matched"
            save_config(cfg)
            st.success("Settings saved to app_config.json")

        # Logout button
        if st.button("Logout"):
            st.session_state["admin_authenticated"] = False
            st.info("🔒 Logged out")
            st.rerun()  # 🔥 force refresh so settings disappear
# -------------------------
# App (User) page
# -------------------------
elif page == "App":
    st.header("User — Upload or Paste input & find nearest backend site")
        # Check user login status
    if not st.session_state["user_authenticated"]:
        user_login()
        st.stop()
    else:
        st.success("✅ Welcome, User!")
            
        
        # Load backend (prefer session_state if admin uploaded one)
      #  if "backend_df" in st.session_state and not st.session_state["backend_df"].empty:
       #     backend_df = st.session_state["backend_df"]
        #    backend_df = normalize_latlon_names(backend_df)
       # else:
        #    if not os.path.exists(BACKEND_FILE):
         #       st.warning("No backend_data.csv found. Ask Admin to upload backend data.")
          #      st.stop()
           # try:
            #    backend_df = safe_read_table(BACKEND_FILE, filename=BACKEND_FILE)
             #   backend_df = normalize_latlon_names(backend_df)
            #except Exception as e:
             #   st.error(f"Failed to load backend: {e}")
              #  st.stop()



    # Load backend
    if not os.path.exists(BACKEND_FILE):
        st.warning("No backend_data.csv found. Ask Admin to upload backend data.")
        st.stop()

    try:
        backend_df = safe_read_table(BACKEND_FILE, filename=BACKEND_FILE)
        backend_df = normalize_latlon_names(backend_df)
    except Exception as e:
        st.error(f"Failed to load backend: {e}")
        st.stop()
        
      #  st.write(f"✅ Loaded {len(backend_df)} backend rows.")   
    # Show backend summary

    st.write(f"**Serving Sites:** {len(backend_df)}")

    # persist user input in session
    if "user_df" not in st.session_state:
        st.session_state["user_df"] = None

    # Input method
    st.subheader("1) Provide your input data")
    tab1, tab2 = st.tabs(["Upload file", "Paste table"])

    with tab1:
        user_file = st.file_uploader("Upload CSV/XLSX input (contains your points)", type=["csv","xlsx"], key="upload_input")
        if user_file is not None:
            try:
                udf = safe_read_table(user_file, filename=getattr(user_file, "name", None))
                udf = normalize_latlon_names(udf)
                st.session_state["user_df"] = udf
                st.success(f"Uploaded input data: {len(udf)} rows")
            except Exception as e:
                st.error(f"Failed to parse uploaded input: {e}")

    with tab2:
        pasted = st.text_area("Paste tab-separated or comma-separated data (include header)", height=200, key="paste_input")
        if st.button("Parse pasted data", key="parse_paste"):
            if not pasted.strip():
                st.warning("Please paste some data first.")
            else:
                try:
                    # prefer tab if present
                    if "\t" in pasted:
                        udf = pd.read_csv(StringIO(pasted), sep="\t", engine="python")
                    elif "," in pasted:
                        udf = pd.read_csv(StringIO(pasted), sep=",", engine="python")
                    else:
                        udf = pd.read_csv(StringIO(pasted), delim_whitespace=True, engine="python")
                    udf = normalize_latlon_names(udf)
                    st.session_state["user_df"] = udf
                    st.success(f"Parsed pasted input: {len(udf)} rows")
                except Exception as e:
                    st.error(f"Failed to parse pasted data: {e}")

    # clear
    if st.session_state.get("user_df") is not None:
        if st.button("Clear uploaded/pasted input"):
            st.session_state["user_df"] = None
            st.rerun()

    # if no input, stop
    user_df = st.session_state.get("user_df")
    if user_df is None or user_df.empty:
        st.info("Upload or paste your input to continue.")
        st.stop()

    # preview
    st.subheader("Preview input (first 10 rows)")
    st.dataframe(user_df.head(10))

    # Lat/Lon detection & selection
    st.subheader("2) Latitude / Longitude selection")
    guessed_lat, guessed_lon = detect_latlon_candidates(user_df)
    cols = list(user_df.columns)
    lat_index = cols.index(guessed_lat) if guessed_lat in cols else (0 if len(cols)>0 else None)
    lon_index = cols.index(guessed_lon) if guessed_lon in cols else (1 if len(cols)>1 else None)
    col1, col2 = st.columns(2)
    with col1:
        user_lat_col = st.selectbox("Latitude column (choose)", options=cols, index=lat_index)
    with col2:
        user_lon_col = st.selectbox("Longitude column (choose)", options=cols, index=lon_index)

    # Coordinate format choices
    fmt = st.selectbox("Coordinate format", ["Decimal degrees", "Scaled integers (E7)", "Swap lat/lon"])
    auto_swap = st.checkbox("Auto-detect swap if lat/lon look reversed", value=False)

    # apply format adjustments for matching
    def apply_coord_format(df, latc, lonc, fmt_key, auto_swap=False):
        df2 = df.copy()
        df2[latc] = pd.to_numeric(df2[latc], errors="coerce")
        df2[lonc] = pd.to_numeric(df2[lonc], errors="coerce")
        if fmt_key == "Scaled integers (E7)":
            df2[latc] = df2[latc] / 1e7
            df2[lonc] = df2[lonc] / 1e7
        if fmt_key == "Swap lat/lon":
            df2[latc], df2[lonc] = df2[lonc], df2[latc]
        if auto_swap:
            # basic heuristic: if many lat values out of -90..90, swap
            lat = df2[latc]
            lon = df2[lonc]
            lat_bad = ((lat < -90) | (lat > 90)).mean() if len(lat) else 0
            lon_ok_for_lat = ((lon >= -90) & (lon <= 90)).mean() if len(lon) else 0
            if lat_bad > 0.5 and lon_ok_for_lat > 0.5:
                df2[latc], df2[lonc] = df2[lonc], df2[latc]
        return df2

    user_df_fixed = apply_coord_format(user_df, user_lat_col, user_lon_col, fmt, auto_swap=auto_swap)

    # show small quality summary
    with st.expander("Coordinate quality summary"):
        q = {
            "rows": len(user_df_fixed),
            "nan_lat": int(pd.to_numeric(user_df_fixed[user_lat_col], errors="coerce").isna().sum()),
            "nan_lon": int(pd.to_numeric(user_df_fixed[user_lon_col], errors="coerce").isna().sum())
        }
        st.json(q)
        if q["nan_lat"] or q["nan_lon"]:
            st.warning("Some coordinates are missing or invalid. Check your selection/formatting.")

    # Filters
    st.subheader("3) Filters (optional)")
    # default: show only non-lat/lon columns from backend as filterable
    filter_candidates = [c for c in backend_df.columns if c not in ["Latitude","Longitude"]]
    filter_cols = st.multiselect("Choose backend columns to filter", options=filter_candidates, default=[])
    backend_filtered = backend_df.copy()
    active_filters = {}
    for col in filter_cols:
        vals = sorted(backend_df[col].dropna().astype(str).unique().tolist())
        sel = st.multiselect(f"Values for {col}", options=vals, default=None, key=f"filter_{col}")
        if sel:
            backend_filtered = backend_filtered[backend_filtered[col].astype(str).isin(sel)]
            active_filters[col] = sel

    if active_filters:
        st.markdown("**Active filters:**")
        for k,v in active_filters.items():
            st.write(f"- **{k}**: {', '.join(v)}")

    st.write(f"Serving Sites after filters: {len(backend_filtered)}")
    if backend_filtered.empty:
        st.error("No Serving Site left after filters.")
        st.stop()

    # Prepare backend coords for fast distance calc
    backend_filtered["Latitude"] = pd.to_numeric(backend_filtered["Latitude"], errors="coerce")
    backend_filtered["Longitude"] = pd.to_numeric(backend_filtered["Longitude"], errors="coerce")
    backend_filtered = backend_filtered.dropna(subset=["Latitude","Longitude"]).reset_index(drop=True)
    backend_lat_rad = np.radians(backend_filtered["Latitude"].to_numpy(dtype=float))
    backend_lon_rad = np.radians(backend_filtered["Longitude"].to_numpy(dtype=float))

    # 4) Matching options
    st.subheader("4) Matching options")
    global_nth = st.number_input("Global Nth nearest (used when no per-row override)", min_value=1, value=1, step=1)
    distance_unit = st.radio("Distance unit", ("Kilometers","Miles"), horizontal=True)
    st.write(f"Feasible threshold (admin-set): **{cfg.get('feasible_km', DEFAULT_CONFIG['feasible_km'])} km**")
    conflict_suffix = cfg.get("backend_conflict_suffix", DEFAULT_CONFIG["backend_conflict_suffix"])

    # Determine if input has per-row nth override column
    nth_col_candidates = [c for c in user_df_fixed.columns if "".join(ch.lower() for ch in c if ch.isalnum()) in ("n","nth","rank","k")]
    per_row_nth_col = nth_col_candidates[0] if nth_col_candidates else None
    if per_row_nth_col:
        st.info(f"Per-row Nth override detected in input column: {per_row_nth_col}")

    # -------------------------
    # Run matching
    # -------------------------
    if st.button("Run matching"):
        # Run matching logic (same as before) - produces `final`
        df_user = user_df_fixed.copy()
        # ensure numeric lat/lon
        df_user[user_lat_col] = pd.to_numeric(df_user[user_lat_col], errors="coerce")
        df_user[user_lon_col] = pd.to_numeric(df_user[user_lon_col], errors="coerce")
        df_user = df_user.dropna(subset=[user_lat_col, user_lon_col]).reset_index(drop=True)
        if df_user.empty:
            st.error("No valid input rows to process.")
            st.stop()

        results_frames = []
        input_columns = list(df_user.columns)

        for idx, row in df_user.iterrows():
            lat_val = float(row[user_lat_col])
            lon_val = float(row[user_lon_col])

            # decide nth for this row
            nth = global_nth
            if per_row_nth_col:
                try:
                    alt = int(pd.to_numeric(row[per_row_nth_col], errors="coerce"))
                    if not np.isnan(alt) and alt >= 1:
                        nth = alt
                except Exception:
                    pass

            dists = vectorized_haversine(lat_val, lon_val, backend_lat_rad, backend_lon_rad)
            j = find_nth_index(dists, n=int(nth))
            if j is None:
                empty_backend = {c: None for c in backend_filtered.columns}
                empty_backend.update({"Distance_km": None, "Distance_miles": None, "Feasible": None})
                backend_row = pd.Series(empty_backend)
            else:
                backend_row = backend_filtered.iloc[int(j)].copy()
                dist_km = float(dists[j])
                backend_row["Distance_km"] = round(dist_km, 6)
                backend_row["Distance_miles"] = round(dist_km * 0.621371, 6)
                backend_row["Feasible"] = "Feasible" if dist_km <= float(cfg.get("feasible_km", DEFAULT_CONFIG["feasible_km"])) else "Not Feasible"

            # Merge input row and backend_row into one DataFrame row
            merge_map = backend_column_merge_dict(backend_row, input_columns, conflict_suffix)
            backend_row_renamed = backend_row.rename(index=merge_map).to_frame().T.reset_index(drop=True)
            input_row_df = row.to_frame().T.reset_index(drop=True)
            combined = pd.concat([input_row_df, backend_row_renamed], axis=1)
            combined["Nth_used"] = int(nth)
            results_frames.append(combined)

        # Final results DataFrame
        final = pd.concat(results_frames, ignore_index=True)

        # Normalize types & rounding
        if "Distance_km" in final.columns:
            final["Distance_km"] = pd.to_numeric(final["Distance_km"], errors="coerce").round(2)
        if "Distance_miles" in final.columns:
            final["Distance_miles"] = pd.to_numeric(final["Distance_miles"], errors="coerce").round(2)

        # Save to session_state for later safe use
        st.session_state["final"] = final

        st.success("Matching completed.")
        # show immediate preview (single preview only)
        #st.subheader("Results (first 200 rows)")
       # st.dataframe(final.head(200), width='stretch')

    # ---------------------------------------
    # After matching: safe preview + Excel export
    # ---------------------------------------
    # Use the results only if they exist in session_state
    if "final" in st.session_state and st.session_state["final"] is not None:
        final = st.session_state["final"]
               # (everything for preview + Excel export stays here)
    else:
        st.empty()  # keeps layout clean, no red error box
            
    # --- Build consistent column groups (updated & robust) ---
    if "final" in st.session_state and isinstance(st.session_state["final"], pd.DataFrame):
        final = st.session_state["final"]
        input_cols = list(user_df.columns)

        # Detect site columns
        site_code_col = next((c for c in final.columns if "site" in c.lower() and "code" in c.lower()), None)
        site_name_col = next((c for c in final.columns if "site" in c.lower() and "name" in c.lower()), None)
        site_cols = [c for c in (site_code_col, site_name_col) if c is not None]

        # Calculated columns
        calc_cols = site_cols + ["Distance_km", "Distance_miles", "Feasible", "Nth_used"]
        calc_cols = [c for c in calc_cols if c in final.columns]

        # Backend columns
        backend_cols = [c for c in final.columns if c not in input_cols + calc_cols]

        # Reorder columns for display
        new_order = input_cols + calc_cols + backend_cols
        final = final.reindex(columns=[c for c in new_order if c in final.columns])

        # --- Grouped preview in Streamlit ---
        group_labels = (
            ["Input Data"] * len(input_cols)
            + ["Calculated"] * len(calc_cols)
            + ["Backend Data"] * len(backend_cols)
        )
        preview_df = final.head(200).copy()
        preview_df.columns = pd.MultiIndex.from_arrays([group_labels, preview_df.columns])

        st.subheader("Results (first 200 rows) — grouped preview")
        st.dataframe(preview_df, width='stretch')

        # --- Build Excel with grouped headers ---
        import io
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.utils import get_column_letter

        wb = Workbook()
        ws = wb.active
        ws.title = "Results"

        # Header styles
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill("solid", fgColor="4F81BD")
        subheader_fill = PatternFill("solid", fgColor="D9E1F2")
        center = Alignment(horizontal="center", vertical="center")

        col_idx = 1
        groups = [
            ("Input Data", input_cols),
            ("Calculated", calc_cols),
            ("Backend Data", backend_cols),
        ]

        for group_name, cols in groups:
            if not cols:
                continue
            start_col = col_idx
            end_col = col_idx + len(cols) - 1

            ws.merge_cells(start_row=1, start_column=start_col, end_row=1, end_column=end_col)
            cell = ws.cell(row=1, column=start_col, value=group_name)
            cell.font = header_font
            cell.alignment = center
            cell.fill = header_fill

            for i, col in enumerate(cols):
                c = ws.cell(row=2, column=start_col + i, value=str(col))
                c.font = Font(bold=True)
                c.alignment = center
                c.fill = subheader_fill

            col_idx = end_col + 1

        # Write data
        for r_idx, row in enumerate(final.itertuples(index=False, name=None), start=3):
            for c_idx, value in enumerate(row, start=1):
                ws.cell(row=r_idx, column=c_idx, value=value)

        # Adjust widths
        for i, col in enumerate(final.columns, start=1):
            letter = get_column_letter(i)
            ws.column_dimensions[letter].width = min(max(8, len(str(col)) + 2), 40)

        excel_buffer = io.BytesIO()
        wb.save(excel_buffer)
        excel_buffer.seek(0)

        st.download_button(
            "Download results (Excel with grouped headers)",
            data=excel_buffer.getvalue(),
            file_name="nearest_results.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

        if st.button("Logout", key="user_logout"):
            st.session_state["user_authenticated"] = False
            st.info("🔒 Logged out successfully")
            st.rerun()

 #       st.info("Run matching to see results and download Excel.")
        #csv_bytes = final.to_csv(index=False).encode("utf-8")
        #st.download_button("Download results CSV", data=csv_bytes, file_name="nearest_results.csv", mime="text/csv")
